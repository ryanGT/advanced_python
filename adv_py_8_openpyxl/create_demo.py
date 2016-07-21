from openpyxl import Workbook
# adapted from the openpyxl docs:
# http://openpyxl.readthedocs.io/en/default/usage.html#read-an-existing-workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from numpy import column_stack, arange, pi, sin

wb = Workbook()
dest_filename = 'created_book.xlsx'

ws1 = wb.active
ws1.title = "saved array"

labels = ['t','y']
ws1.append(labels)

t = arange(0,1,0.01)
y = sin(2*pi*t)

data_out = column_stack([t,y])

for row in data_out.tolist():
     ws1.append(row)

ws2 = wb.create_sheet(title="hello")

ws2['B5'] = 1.1234
ws2['C3'] = 'hello'

wb.save(filename = dest_filename)
